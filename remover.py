import argparse
from argparse import ArgumentParser, Namespace

from openpyxl.reader.excel import load_workbook


parser = ArgumentParser()
parser.add_argument('file_name', type=str, help='Excel file name')
parser.add_argument('--column', dest='column', default='A', help='Specify which column to check')
parser.add_argument('--skip-row', dest='skip_rows', type=int, default=1, help='Skip first n rows')
parser.add_argument('--result', dest='result_name', default=None, help='Specify result file name')


def main(args: Namespace):
    def get_result_name(file_name: str):
        return file_name.replace('.xlsx', '_new.xlsx')

    def remove_rows(ws, info_list: list):
        to_minus = 0
        for info in info_list:
            ws.delete_rows(info['index'] - to_minus, info['count'])
            to_minus += info['count']

    def find_duplication(ws, col: int = 0, skip_row: int = 1):
        rows_to_remove = []

        count = -1
        start_row = 1 + skip_row
        rows = ws.iter_rows(min_row=start_row)
        for index, row in enumerate(rows):
            if count == -1:
                count = 0
                continue

            last_row = ws[index + skip_row]
            if last_row[col].value != row[col].value:
                if count > 0:
                    # remove duplication
                    rows_to_remove.append({
                        'index': index - count + skip_row,
                        'count': count
                    })
                    # reset count
                    count = 0
                    print(f'Found \'{last_row[col].value}\'')
            else:
                count += 1

            if index == 50:
                break

        return rows_to_remove

    # parse arguments
    file_name = args.file_name
    result_name = get_result_name(file_name) if args.result_name is None else args.result_name
    skip_rows = args.skip_rows
    if len(args.column) == 1:  # TODO: support for column AA-AZ...
        column = ord(args.column) - ord('A')
    else:
        print('Column name length must be 1')
        return

    # laod worksheet
    print(f'Loading file {file_name}...')
    wb = load_workbook(filename=file_name)
    ws = wb.active
    print(f'Finding duplication in column {args.column}...\n')
    rows_to_remove = find_duplication(ws, col=column, skip_row=skip_rows)
    print(f'\nReducing {len(rows_to_remove)} groups...')
    remove_rows(ws, info_list=rows_to_remove)
    print(f'Saving result file to {result_name}...')
    wb.save(result_name)


if __name__ == '__main__':
    args = parser.parse_args()
    main(args)
